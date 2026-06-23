#!/usr/bin/env python3
"""
收货登记 Excel 管理工具

支持子命令：
  init        初始化收货登记Excel文件
  next-number 获取下一个收货记录编号
  add         添加收货记录（支持多物品）
  query       查询收货记录
  update      更新收货记录

依赖：openpyxl
用法示例见 SKILL.md
"""

import argparse
import json
import sys
import os
from datetime import datetime, date

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print(json.dumps({
        "success": False,
        "error": "缺少依赖库 openpyxl，请运行: pip install openpyxl"
    }, ensure_ascii=False))
    sys.exit(1)


# ============================================================
# 常量定义
# ============================================================

# Excel 表头（第1行）
HEADERS = [
    "收货记录编号", "收货日期", "供应商名称", "采购订单号",
    "物流/快递单号", "收货人", "物品名称", "规格型号",
    "数量", "单位", "批次号", "生产日期",
    "有效期至", "序列号", "质检状态", "项目归属",
    "登记时间", "备注"
]

# 列宽配置（列字母: 宽度）
COLUMN_WIDTHS = {
    'A': 22, 'B': 14, 'C': 28, 'D': 18,
    'E': 22, 'F': 12, 'G': 28, 'H': 18,
    'I': 10, 'J': 8,  'K': 18, 'L': 14,
    'M': 14, 'N': 22, 'O': 12, 'P': 28,
    'Q': 22, 'R': 30
}

# 表头样式
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 数据行样式
DATA_FONT = Font(name="微软雅黑", size=10)
DATA_ALIGNMENT = Alignment(vertical="center", wrap_text=True)

# 边框
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# 物品明细字段（每个物品独立填写）
ITEM_FIELDS = [
    "物品名称", "规格型号", "数量", "单位",
    "批次号", "生产日期", "有效期至", "序列号",
    "质检状态"
]

# 公共信息字段（一次收货共享）
COMMON_FIELDS = [
    "收货日期", "供应商名称", "采购订单号",
    "物流/快递单号", "收货人", "项目归属", "备注"
]

# 系统自动生成字段
AUTO_FIELDS = ["收货记录编号", "登记时间"]

# 全部字段顺序（与 HEADERS 一致）
ALL_FIELDS = HEADERS[:]


# ============================================================
# 工具函数
# ============================================================

def output_json(data):
    """输出 JSON 到 stdout"""
    print(json.dumps(data, ensure_ascii=False, default=str))


def ensure_dir(file_path):
    """确保文件所在目录存在"""
    dir_path = os.path.dirname(os.path.abspath(file_path))
    if dir_path and not os.path.exists(dir_path):
        os.makedirs(dir_path, exist_ok=True)


def parse_date(date_str):
    """解析日期字符串，返回 YYYY-MM-DD 格式"""
    if not date_str:
        return date.today().strftime("%Y-%m-%d")
    try:
        # 尝试解析常见格式
        for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y%m%d", "%Y年%m月%d日"]:
            try:
                return datetime.strptime(date_str, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
        return date_str
    except Exception:
        return date_str


# ============================================================
# init - 初始化 Excel 文件
# ============================================================

def cmd_init(args):
    """初始化收货登记 Excel 文件"""
    file_path = args.file
    company_name = args.company

    if not company_name:
        output_json({"success": False, "error": "必须指定主体名称 (--company)"})
        return

    # 检查文件是否已存在
    if os.path.exists(file_path):
        output_json({
            "success": False,
            "error": f"文件已存在: {file_path}，如需重新初始化请先删除或重命名原文件"
        })
        return

    ensure_dir(file_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "收货登记"

    # 写入表头
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # 设置列宽
    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # 冻结首行
    ws.freeze_panes = "A2"

    # 设置行高
    ws.row_dimensions[1].height = 30

    # 添加自动筛选
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    wb.save(file_path)

    output_json({
        "success": True,
        "message": "收货登记表初始化成功",
        "file_path": file_path,
        "company": company_name,
        "headers": HEADERS,
        "header_count": len(HEADERS)
    })


# ============================================================
# next-number - 获取下一个收货记录编号
# ============================================================

def cmd_next_number(args):
    """获取下一个收货记录编号"""
    file_path = args.file
    date_str = parse_date(args.date)

    if not os.path.exists(file_path):
        output_json({"success": False, "error": f"文件不存在: {file_path}，请先初始化"})
        return

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    # 收货日期在第2列（B列）
    # 收货记录编号在第1列（A列）
    date_compact = date_str.replace("-", "")  # YYYYMMDD
    prefix = f"RG-{date_compact}"

    # 统计当日已有记录数
    count = 0
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=2, values_only=True):
        record_number = row[0]
        receive_date = row[1]
        if record_number is None:
            continue
        # 通过编号前缀匹配
        if isinstance(record_number, str) and record_number.startswith(prefix):
            count += 1
        # 也通过日期匹配（兼容手动输入的情况）
        elif receive_date is not None:
            row_date = str(receive_date).strip()[:10].replace("-", "")
            if row_date == date_compact:
                count += 1

    wb.close()

    next_number = f"RG-{date_compact}-{count + 1:03d}"

    output_json({
        "success": True,
        "next_number": next_number,
        "date": date_str,
        "existing_count": count
    })


# ============================================================
# add - 添加收货记录
# ============================================================

def cmd_add(args):
    """添加收货记录（支持多物品）"""
    file_path = args.file

    if not os.path.exists(file_path):
        output_json({"success": False, "error": f"文件不存在: {file_path}，请先初始化"})
        return

    # 解析 JSON 数据
    try:
        data = json.loads(args.data)
    except json.JSONDecodeError as e:
        output_json({"success": False, "error": f"JSON 解析失败: {e}"})
        return

    common = data.get("common", {})
    items = data.get("items", [])

    if not items:
        output_json({"success": False, "error": "物品明细不能为空"})
        return

    # 校验必填字段
    required_common = ["收货日期", "供应商名称", "收货人"]
    for field in required_common:
        if not common.get(field):
            output_json({"success": False, "error": f"公共信息缺少必填字段: {field}"})
            return

    required_item = ["物品名称", "数量", "单位"]
    for i, item in enumerate(items, 1):
        for field in required_item:
            if not item.get(field) and item.get(field) != 0:
                output_json({"success": False, "error": f"第{i}个物品缺少必填字段: {field}"})
                return

        # 校验数量必须为正数（兼容字符串传入）
        quantity = item.get("数量")
        try:
            quantity_num = float(quantity) if quantity is not None else 0
        except (ValueError, TypeError):
            output_json({"success": False, "error": f"第{i}个物品数量必须为正数，当前值: {quantity}"})
            return
        if quantity_num <= 0:
            output_json({"success": False, "error": f"第{i}个物品数量必须为正数，当前值: {quantity}"})
            return

        # 物品类型条件校验
        item_type = item.get("物品类型", "")
        if item_type == "仪器设备":
            if not item.get("序列号"):
                output_json({"success": False, "error": f"第{i}个物品为仪器设备，序列号为必填字段"})
                return
        else:
            if not item.get("批次号"):
                output_json({"success": False, "error": f"第{i}个物品批次号为必填字段（仪器设备除外）"})
                return

    # 获取收货日期
    receive_date = parse_date(common.get("收货日期"))
    date_compact = receive_date.replace("-", "")
    prefix = f"RG-{date_compact}"

    # 加载工作簿（非只读模式，用于写入）
    wb = load_workbook(file_path)
    ws = wb.active

    # 找到第一个空行
    start_row = ws.max_row + 1
    # 检查该行是否真的为空（max_row 可能有格式但无数据）
    while ws.cell(row=start_row, column=1).value is not None:
        start_row += 1

    # 统计当日已有记录数以生成编号
    existing_count = 0
    for row in range(2, start_row):
        record_number = ws.cell(row=row, column=1).value
        if isinstance(record_number, str) and record_number.startswith(prefix):
            existing_count += 1

    # 当前时间
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # 写入每条记录
    written_records = []
    for i, item in enumerate(items):
        row_num = start_row + i
        record_number = f"RG-{date_compact}-{existing_count + i + 1:03d}"

        # 构建完整记录
        record = {}
        record["收货记录编号"] = record_number
        record["收货日期"] = receive_date
        record["供应商名称"] = common.get("供应商名称", "")
        record["采购订单号"] = common.get("采购订单号", "")
        record["物流/快递单号"] = common.get("物流/快递单号", "")
        record["收货人"] = common.get("收货人", "")
        record["物品名称"] = item.get("物品名称", "")
        record["规格型号"] = item.get("规格型号", "")
        record["数量"] = item.get("数量", "")
        record["单位"] = item.get("单位", "")
        record["批次号"] = item.get("批次号", "")
        record["生产日期"] = item.get("生产日期", "")
        record["有效期至"] = item.get("有效期至", "")
        record["序列号"] = item.get("序列号", "")
        record["质检状态"] = item.get("质检状态", "待检")
        record["项目归属"] = common.get("项目归属", "")
        record["登记时间"] = now_str
        record["备注"] = common.get("备注", "")

        # 写入行
        for col_idx, field_name in enumerate(HEADERS, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=record[field_name])
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGNMENT
            cell.border = THIN_BORDER

        # 摘要信息（不含所有字段，仅关键信息）
        written_records.append({
            "收货记录编号": record_number,
            "收货日期": receive_date,
            "供应商名称": record["供应商名称"],
            "物品名称": record["物品名称"],
            "数量": record["数量"],
            "单位": record["单位"],
            "登记时间": now_str
        })

    wb.save(file_path)
    wb.close()

    output_json({
        "success": True,
        "message": f"成功写入 {len(written_records)} 条收货记录",
        "count": len(written_records),
        "records": written_records,
        "file_path": file_path
    })


# ============================================================
# query - 查询收货记录
# ============================================================

def cmd_query(args):
    """查询收货记录"""
    file_path = args.file

    if not os.path.exists(file_path):
        output_json({"success": False, "error": f"文件不存在: {file_path}"})
        return

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    results = []

    # 构建筛选条件
    filters = {}
    if args.number:
        filters["收货记录编号"] = args.number
    if args.date:
        filters["收货日期"] = parse_date(args.date)
    if args.supplier:
        filters["供应商名称"] = args.supplier

    for row in ws.iter_rows(min_row=2, values_only=True):
        # 跳过空行
        if all(cell is None for cell in row):
            continue

        record = {}
        for col_idx, header in enumerate(HEADERS):
            value = row[col_idx] if col_idx < len(row) else None
            if value is not None:
                record[header] = str(value) if not isinstance(value, (int, float)) else value
            else:
                record[header] = ""

        # 应用筛选
        match = True
        for field, filter_value in filters.items():
            record_value = str(record.get(field, ""))
            if field == "收货记录编号":
                # 编号字段使用精确匹配
                if record_value != filter_value:
                    match = False
                    break
            else:
                # 其他字段使用模糊匹配
                if filter_value not in record_value:
                    match = False
                    break

        if match:
            results.append(record)

    wb.close()

    output_json({
        "success": True,
        "count": len(results),
        "records": results
    })


# ============================================================
# update - 更新收货记录
# ============================================================

def cmd_update(args):
    """更新收货记录"""
    file_path = args.file
    record_number = args.number

    if not os.path.exists(file_path):
        output_json({"success": False, "error": f"文件不存在: {file_path}"})
        return

    # 解析更新数据
    try:
        updates = json.loads(args.data)
    except json.JSONDecodeError as e:
        output_json({"success": False, "error": f"JSON 解析失败: {e}"})
        return

    wb = load_workbook(file_path)
    ws = wb.active

    # 查找目标行
    target_row = None
    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value and str(cell_value).strip() == record_number:
            target_row = row
            break

    if target_row is None:
        wb.close()
        output_json({"success": False, "error": f"未找到编号为 {record_number} 的记录"})
        return

    # 读取更新前的记录
    before = {}
    for col_idx, header in enumerate(HEADERS, 1):
        value = ws.cell(row=target_row, column=col_idx).value
        before[header] = value

    # 应用更新
    changed_fields = []
    for field, new_value in updates.items():
        if field in HEADERS:
            col_idx = HEADERS.index(field) + 1
            old_value = ws.cell(row=target_row, column=col_idx).value
            if str(old_value) != str(new_value):
                changed_fields.append({
                    "field": field,
                    "old_value": old_value,
                    "new_value": new_value
                })
            cell = ws.cell(row=target_row, column=col_idx, value=new_value)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGNMENT
            cell.border = THIN_BORDER

    wb.save(file_path)

    # 读取更新后的完整记录
    after = {}
    for col_idx, header in enumerate(HEADERS, 1):
        after[header] = ws.cell(row=target_row, column=col_idx).value

    wb.close()

    output_json({
        "success": True,
        "message": f"记录 {record_number} 已更新",
        "record_number": record_number,
        "changed_fields": changed_fields,
        "changed_count": len(changed_fields),
        "before": before,
        "after": after
    })


# ============================================================
# list-companies - 列出已配置的主体（从 company_registry.md 解析）
# ============================================================

def cmd_list_companies(args):
    """列出已配置的主体列表"""
    registry_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "references", "company_registry.md"
    )

    companies = []
    if os.path.exists(registry_path):
        with open(registry_path, 'r', encoding='utf-8') as f:
            content = f.read()
        # 简单解析表格行
        for line in content.split('\n'):
            line = line.strip()
            if line.startswith('|') and not line.startswith('|---') and not line.startswith('| 序号') and not line.startswith('| 序号'):
                parts = [p.strip() for p in line.split('|')]
                if len(parts) >= 5 and parts[1].isdigit():
                    full_name = parts[2]
                    short_name = parts[3]
                    # 排除占位项（含方括号）
                    if '[' not in full_name and full_name:
                        companies.append({
                            "full_name": full_name,
                            "short_name": short_name,
                            "file_name": parts[4]
                        })

    output_json({
        "success": True,
        "companies": companies,
        "count": len(companies)
    })


# ============================================================
# 主入口
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description="收货登记 Excel 管理工具",
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    subparsers = parser.add_subparsers(dest="command", help="子命令")

    # init
    p_init = subparsers.add_parser("init", help="初始化收货登记Excel文件")
    p_init.add_argument("--file", required=True, help="Excel 文件路径")
    p_init.add_argument("--company", required=True, help="主体全称")

    # next-number
    p_num = subparsers.add_parser("next-number", help="获取下一个收货记录编号")
    p_num.add_argument("--file", required=True, help="Excel 文件路径")
    p_num.add_argument("--date", default=None, help="收货日期（YYYY-MM-DD），默认当天")

    # add
    p_add = subparsers.add_parser("add", help="添加收货记录")
    p_add.add_argument("--file", required=True, help="Excel 文件路径")
    p_add.add_argument("--data", required=True, help="JSON 格式的收货数据")

    # query
    p_query = subparsers.add_parser("query", help="查询收货记录")
    p_query.add_argument("--file", required=True, help="Excel 文件路径")
    p_query.add_argument("--number", default=None, help="按收货记录编号查询")
    p_query.add_argument("--date", default=None, help="按收货日期查询")
    p_query.add_argument("--supplier", default=None, help="按供应商名称查询（模糊匹配）")

    # update
    p_update = subparsers.add_parser("update", help="更新收货记录")
    p_update.add_argument("--file", required=True, help="Excel 文件路径")
    p_update.add_argument("--number", required=True, help="收货记录编号")
    p_update.add_argument("--data", required=True, help="JSON 格式的更新数据")

    # list-companies
    p_list = subparsers.add_parser("list-companies", help="列出已配置的主体")

    args = parser.parse_args()

    if args.command == "init":
        cmd_init(args)
    elif args.command == "next-number":
        cmd_next_number(args)
    elif args.command == "add":
        cmd_add(args)
    elif args.command == "query":
        cmd_query(args)
    elif args.command == "update":
        cmd_update(args)
    elif args.command == "list-companies":
        cmd_list_companies(args)
    else:
        parser.print_help()
        sys.exit(1)


if __name__ == '__main__':
    main()
